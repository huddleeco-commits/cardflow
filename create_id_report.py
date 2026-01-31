import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from PIL import Image as PILImage
import os
from datetime import datetime

# Card identification data
cards = [
    {
        "player": "Victor Wembanyama",
        "year": 2023,
        "set": "Panini Obsidian",
        "card_num": "9",
        "parallel": "Electric Etch Neon Flood",
        "numbered": "SSP (unnumbered)",
        "team": "San Antonio Spurs",
        "sport": "Basketball",
        "graded": "Yes",
        "grading_company": "PSA",
        "grade": "9",
        "condition": "Mint",
        "confidence": "High",
        "notes": "Rookie Card (RC). Neon Flood is a hobby-exclusive SSP parallel.",
        "front_img": "93779221-1.jpg",
        "back_img": "93779221-2.jpg"
    },
    {
        "player": "Shohei Ohtani",
        "year": 2021,
        "set": "Topps Heritage",
        "card_num": "245",
        "parallel": "Base",
        "numbered": "N/A",
        "team": "Los Angeles Angels",
        "sport": "Baseball",
        "graded": "No",
        "grading_company": "N/A",
        "grade": "N/A",
        "condition": "Near Mint",
        "confidence": "High",
        "notes": "1972 Topps retro design. Card shows Ohtani in batting stance.",
        "front_img": "card_card_1768355971105_a.jpg",
        "back_img": "card_card_1768355971105_b.jpg"
    },
    {
        "player": "Bulbasaur",
        "year": 2023,
        "set": "Pokemon 151 (Scarlet & Violet)",
        "card_num": "166/165",
        "parallel": "Special Art Rare (Illustration Rare)",
        "numbered": "N/A",
        "team": "N/A",
        "sport": "Pokemon",
        "graded": "No",
        "grading_company": "N/A",
        "grade": "N/A",
        "condition": "Near Mint",
        "confidence": "High",
        "notes": "Full art illustration card. Part of the Pokemon 151 special set celebrating original 151 Pokemon.",
        "front_img": "poketest2.png",
        "back_img": "poketest2b.jpg"
    }
]

base_path = "C:/Users/huddl/OneDrive/Desktop/card-uploads"

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Card Identification Report"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Title row
ws.merge_cells('A1:P1')
title_cell = ws['A1']
title_cell.value = f"Card Identification Report - Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Headers (row 2)
headers = ["#", "Front", "Back", "Player", "Year", "Set", "Card #", "Parallel", "Numbered",
           "Team", "Sport", "Graded", "Grading Company", "Grade", "Condition", "Confidence", "Notes"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center_align

# Column widths
col_widths = {
    'A': 4,   # #
    'B': 12,  # Front
    'C': 12,  # Back
    'D': 20,  # Player
    'E': 6,   # Year
    'F': 28,  # Set
    'G': 10,  # Card #
    'H': 28,  # Parallel
    'I': 18,  # Numbered
    'J': 20,  # Team
    'K': 10,  # Sport
    'L': 8,   # Graded
    'M': 16,  # Grading Company
    'N': 8,   # Grade
    'O': 12,  # Condition
    'P': 10,  # Confidence
    'Q': 50   # Notes
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Create thumbnails
thumb_files = []
for idx, card in enumerate(cards, 1):
    for img_type, img_key in [('front', 'front_img'), ('back', 'back_img')]:
        img_path = os.path.join(base_path, card[img_key])
        if os.path.exists(img_path):
            pil_img = PILImage.open(img_path)
            pil_img.thumbnail((80, 80))
            thumb_path = os.path.join(base_path, f"id_thumb_{idx}_{img_type}.png")
            pil_img.save(thumb_path)
            thumb_files.append(thumb_path)

# Add card data (starting at row 3)
for idx, card in enumerate(cards, 1):
    row = idx + 2
    ws.row_dimensions[row].height = 70

    # Column A: #
    ws.cell(row=row, column=1, value=idx).alignment = center_align
    ws.cell(row=row, column=1).border = border

    # Column B: Front (placeholder)
    ws.cell(row=row, column=2, value="").border = border

    # Column C: Back (placeholder)
    ws.cell(row=row, column=3, value="").border = border

    # Column D: Player
    ws.cell(row=row, column=4, value=card["player"]).alignment = center_align
    ws.cell(row=row, column=4).border = border

    # Column E: Year
    ws.cell(row=row, column=5, value=card["year"]).alignment = center_align
    ws.cell(row=row, column=5).border = border

    # Column F: Set
    ws.cell(row=row, column=6, value=card["set"]).alignment = center_align
    ws.cell(row=row, column=6).border = border

    # Column G: Card #
    ws.cell(row=row, column=7, value=card["card_num"]).alignment = center_align
    ws.cell(row=row, column=7).border = border

    # Column H: Parallel
    ws.cell(row=row, column=8, value=card["parallel"]).alignment = center_align
    ws.cell(row=row, column=8).border = border

    # Column I: Numbered
    ws.cell(row=row, column=9, value=card["numbered"]).alignment = center_align
    ws.cell(row=row, column=9).border = border

    # Column J: Team
    ws.cell(row=row, column=10, value=card["team"]).alignment = center_align
    ws.cell(row=row, column=10).border = border

    # Column K: Sport
    ws.cell(row=row, column=11, value=card["sport"]).alignment = center_align
    ws.cell(row=row, column=11).border = border

    # Column L: Graded
    ws.cell(row=row, column=12, value=card["graded"]).alignment = center_align
    ws.cell(row=row, column=12).border = border

    # Column M: Grading Company
    ws.cell(row=row, column=13, value=card["grading_company"]).alignment = center_align
    ws.cell(row=row, column=13).border = border

    # Column N: Grade
    ws.cell(row=row, column=14, value=card["grade"]).alignment = center_align
    ws.cell(row=row, column=14).border = border

    # Column O: Condition
    ws.cell(row=row, column=15, value=card["condition"]).alignment = center_align
    ws.cell(row=row, column=15).border = border

    # Column P: Confidence
    ws.cell(row=row, column=16, value=card["confidence"]).alignment = center_align
    ws.cell(row=row, column=16).border = border

    # Column Q: Notes
    ws.cell(row=row, column=17, value=card["notes"]).alignment = left_align
    ws.cell(row=row, column=17).border = border

    # Add front thumbnail
    front_thumb = os.path.join(base_path, f"id_thumb_{idx}_front.png")
    if os.path.exists(front_thumb):
        img = Image(front_thumb)
        img.anchor = f'B{row}'
        ws.add_image(img)

    # Add back thumbnail
    back_thumb = os.path.join(base_path, f"id_thumb_{idx}_back.png")
    if os.path.exists(back_thumb):
        img = Image(back_thumb)
        img.anchor = f'C{row}'
        ws.add_image(img)

# Save
output_path = os.path.join(base_path, "identification-report.xlsx")
wb.save(output_path)
print(f"Report saved to: {output_path}")
print(f"Cards identified: {len(cards)}")

# Clean up thumbnails
for thumb_path in thumb_files:
    if os.path.exists(thumb_path):
        os.remove(thumb_path)

print("Thumbnails cleaned up. Done!")
